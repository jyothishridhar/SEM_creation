import streamlit as st
import pandas as pd
import os
import base64
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
import re
from itertools import permutations
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from requests.exceptions import Timeout
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from io import StringIO
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import sqlite3
from io import BytesIO
from openpyxl.styles import PatternFill, Font
import streamlit.components.v1 as components

# Initialize session state
if 'logged_in' not in st.session_state:
    st.session_state.logged_in = False  

# Database setup
def init_db():
    with sqlite3.connect('users.db') as conn:
        c = conn.cursor()
        c.execute('''
            CREATE TABLE IF NOT EXISTS users
            (id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password TEXT NOT NULL)
        ''')
        conn.commit()

        # Insert sample users (for demonstration purposes; ideally use a registration process)
        sample_users = [
            ('admin', 'neona@0625'),
            ('gsadmin', 'Welcome@0625'),
            ('jyo','Rocky@000'),
            # Add more users here
        ]
        
        for user in sample_users:
            c.execute('INSERT OR IGNORE INTO users (username, password) VALUES (?, ?)', user)
        conn.commit()

# Initialize database
init_db()

# Function to get a new connection
def get_db_connection():
    conn = sqlite3.connect('users.db')
    return conn

# Function to verify login credentials
def verify_login(username, password):
    with get_db_connection() as conn:
        c = conn.cursor()
        c.execute('SELECT * FROM users WHERE username=? AND password=?', (username, password))
        return c.fetchone()

# Function to add a new user (optional for registration)
def add_user(username, password):
    with get_db_connection() as conn:
        c = conn.cursor()
        c.execute('INSERT INTO users (username, password) VALUES (?, ?)', (username, password))
        conn.commit()

# Logout function
def logout():
    st.session_state.logged_in = False
    st.success("Logged out successfully!")

# Login function
def login():
    st.title("Login")
    username = st.text_input("Username")
    password = st.text_input("Password", type='password')
    
    if st.button("Login"):
        user = verify_login(username, password)
        if user:
            st.success("Logged in successfully!")
            st.session_state.logged_in = True  
            return True
        else:
            st.error("Invalid username or password")
            return False

        
      

headers = {
    
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
}

def generate_variants(property_name, max_variants=5):
    # Split the property name into words
    words = property_name.split()
    
    # Remove specific words that should not be at the end or have "&" at the end
    words = [word.rstrip('&') for word in words if word.lower() not in ['the', 'and'] and not word.endswith('&')]
    
    # Generate permutations of words
    word_permutations = permutations(words)
    
    # Join permutations to form variant names
    variants = [' '.join(perm) for i, perm in enumerate(word_permutations) if i < max_variants]
    
    return variants

# Define function to scrape the first proper paragraph
def scrape_first_proper_paragraph(url, retries=3, wait_time=10):
    try:
        options = Options()
        options.add_argument('--no-sandbox')
        options.add_argument('--window-size=1420,1080')
        options.add_argument('--headless')
        options.add_argument('--disable-gpu')
        chrome_driver_path="C:\\Users\\Jyothi.S\\AppData\\Local\\chromedriver-win64\\chromedriver.exe"

        service = Service(chrome_driver_path)
        driver = webdriver.Chrome(service=service, options=options)
       
        for attempt in range(retries):
            try:
                driver.get(url)

                time.sleep(10)

                #  Handle cookie consent button
                try:
                    WebDriverWait(driver, wait_time).until(
                        EC.presence_of_element_located((By.XPATH, '//button[contains(text(), "Accept All") or contains(text(), "Allow All") or contains(text(), "Agree All" or contains(text(), "Accept and close")]'))
                    )
                    consent_button = driver.find_element(By.XPATH, '//button[contains(text(), "Accept All") or contains(text(), "Allow All") or contains(text(), "Agree All")]')
                    consent_button.click()
                    print("Cookie consent button clicked.")
                except Exception as e:
                    print(f"No cookie consent button found or failed to click: {e}")
                
               
                # Use explicit wait to ensure the page has fully rendered
                WebDriverWait(driver, wait_time).until(
                    EC.presence_of_element_located((By.TAG_NAME, 'p'))
                )
               
                # Get the page source
                page_source = driver.page_source
               
                # Parse the HTML with BeautifulSoup
                soup = BeautifulSoup(page_source, 'html.parser')
                print("Soup object fetched successfully.")  # Log to check if fetching is successful
               
                # Find all <p> tags
                p_tags = soup.find_all('p')
                # print("Found <p> tags: p_tags",p_tags)
               
                # Initialize a variable to store the text of the first two paragraphs
                first_two_paragraphs_text = ''
                paragraph_count = 0
                seen_paragraphs = set() 

                # Define keywords to exclude
                exclusion_keywords = ['cookie', 'privacy', 'consent', 'policy', 'advertising', 'tracking']
               
                # Find the text of the first two proper paragraphs
                for p in p_tags:
                    paragraph = p.text.strip()
                    print(f"Paragraph {paragraph_count + 1}: {paragraph[:100]}...")  # Print the first 100 characters
                    if len(paragraph) > 50 and not any(keyword in paragraph.lower() for keyword in exclusion_keywords):
                        first_two_paragraphs_text += paragraph + ' '  
                        seen_paragraphs.add(paragraph)
                        paragraph_count += 1
                        print(f"Added Paragraph {paragraph_count}: {paragraph[:100]}...")
                        if paragraph_count == 4:  
                            break

                if paragraph_count < 4 and paragraph_count > 0:
                    print("Less than four proper paragraphs found, but adding the available ones.")
                elif paragraph_count == 0:
                    raise ValueError("No proper paragraphs found.")

                # Log the final text before splitting into sentences
                print(f"First two proper paragraphs text: {first_two_paragraphs_text}")

                # Split the text of the first two paragraphs into sentences
                sentences = re.split(r'(?<!\w\.\w.)(?<![A-Z][a-z]\.)(?<=\.|\?)\s', first_two_paragraphs_text)

                # Remove duplicate sentences
                seen_sentences = set()
                unique_sentences = []
                for sentence in sentences:
                    if sentence not in seen_sentences:
                        unique_sentences.append(sentence)
                        seen_sentences.add(sentence)

                # Log the sentences to debug
                print(f"Unique Sentences: {unique_sentences}")

                # Ensure we have at least eight sentences
                while len(unique_sentences) < 8:
                    unique_sentences.append('')  # Append empty strings if necessary

                return unique_sentences[0] + ' ' + unique_sentences[1] + ' ' +unique_sentences[2] + ' ' + unique_sentences[3], unique_sentences[4] + ' ' + unique_sentences[5] + ' ' +unique_sentences[6] + ' ' + unique_sentences[7]

            except Exception as e:
                print(f"Attempt {attempt + 1} failed with error: {e}")
                time.sleep(5)  # Wait before retrying

        print("All attempts failed. Unable to scrape the paragraphs.")
        return None, None, None, None
    
    finally:
        driver.quit()

def extract_header_from_path(output_file):
    try:
        # Extract filename from the path
        filename = os.path.basename(output_file)
        # Remove extension from filename
        filename_without_extension = os.path.splitext(filename)[0]
        # Replace underscores with spaces
        header_text = filename_without_extension.replace('_', ' ')

        return header_text.strip()

    except Exception as e:
        print("An error occurred while extracting header from file path:", e)
        return None


def scrape_site_links(url, max_links=8):
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
        }

        response = requests.get(url, headers=headers, timeout=15)
        response.raise_for_status()

        # Parse the HTML content
        soup = BeautifulSoup(response.text, 'html.parser')

        # Find the main content and footer sections
        main_content = soup.body
        footer_content = soup.find('footer')

        # Combine all anchor tags from main content and footer
        anchor_tags = main_content.find_all('a') + (footer_content.find_all('a') if footer_content else [])

        # Set to store unique URLs
        unique_urls = set()

        # List to store the found site links
        site_links = []

        # Define patterns to match variations in link text
        link_text_patterns = [
            "Official Site", "Rooms & Suites", "Wedding", "Facilities & Activities", "Sports & Entertainment", 
            "Specials", "Live music", "Stand-up comedy", "Magic shows", "Art exhibitions", "Poolside", "Pool area", 
            "Pool deck", "Pool bar", "Tours & Activities", "All Dining & Bar Facilities", "Activities", "Groups & Meetings", 
            "Dining", "Meetings & Events", "Contact Us", "CONTACT", "EXPERIENCE", "Photos", "Events", "Pool & sea", "Wellness & fitness", 
            "Water Park", "Salt Water Swimming Pool", "Accommodation", "Amenities", "Location", "Rooms", "Gallery", 
            "Pool bar", "Restaurants", "Discover", "Our Services", "Eatery", "Pub", "Diner", "Trattoria", "Brasserie", 
            "Café", "Bistro", "Destination & Location", "Address", "Venue", "Spot", "Place", "Site", "Locale", "Area", 
            "Premises", "Establishment", "Guest Rooms", "Suites", "Deluxe Rooms", "Executive Suites", "Presidential Suite", 
            "Penthouse", "Family Suites", "Connecting Rooms", "Private Suites", "Offers","Our Resorts","Vacation Specials","Park Tickets","Meets us","Stay with us", "Find us","Instagram",
            "Our Hotel","Location","Promotions"
        ]

        # Relevant words related to specific categories
        relevant_meetings_words = ["Meetings & Events", "Groups & Meetings", "Meetings", "Events", "Wedding"]
        relevant_Entertainment_words = ["Sports & Entertainment", "Live music", "Stand-up comedy", "Magic shows", "Art exhibitions", "Sports", "Entertainment"]
        relevant_Facilities_Activities_words = ["Facilities & Activities", "Activities", "Pool & sea", "Salt Water Swimming Pool", "Our Services", "swimming pool", "pool", "sea", "Water Park",  "Poolside", "Pool area", "Pool deck", "Pool bar", "Tours & Activities"]
        relevant_Spa_Wellness_words = ["Spa & Wellness", "Spa", "Wellness & fitness", "Discover"]
        relevant_Photo_Gallery_words = ["PhotoGallery", "Photo", "Gallery"]
        relevant_Dining_words = ["All Dining & Bar Facilities", "Restaurant", "Food & Beverage Amenities", "Dining", "Gastronomy", "Eatery", "Pub", "Diner", "Trattoria", "Brasserie", "Café", "Bistro", "In Room dining", "Private Dining"]
        relevant_Location_words = ["Location", "Locations", "Destination & Location", "Address", "Venue", "Spot", "Place", "Site", "Locale", "Area", "Premises", "Establishment"]
        relevant_Rooms_words = ["Rooms", "Room", "Rooms & Suites", "Rooms and Suites", "Guest Rooms", "Suites", "Deluxe Rooms", "Executive Suites", "Presidential Suite", "Penthouse", "Family Suites", "Connecting Rooms", "Private Suites"]
        relevant_special_offer_words = ["special_offer", "offers", "offer", "Specials","Vacation Specials"]
        relevant_Accommodation_words = ["Explore All Accommodations", "Accommodation", "stay"]
        relevant_specials_packages_words = ["Daily Specials", "Weekend Specials", "Holiday Specials", "Promotional Specials", "Family Packages", "Couples Packages", "Party Packages", "Event Packages", "Set Menus"]

        # Compile regex pattern for link text
        link_text_pattern = re.compile('|'.join(link_text_patterns), re.IGNORECASE)

        for a in anchor_tags:
            link_text = a.get_text(strip=True)
            link_url = a.get('href')

            # Check if the URL is valid
            if link_url:
                # If it's a fragment identifier, prepend the base URL
                if link_url.startswith("#"):
                    link_url = url + link_url
                
                # Skip invalid URLs
                if link_url.startswith(("javascript:", "mailto:", "tel:")):
                    continue
                
                link_url = urljoin(url, link_url)

                if link_url not in unique_urls:
                    unique_urls.add(link_url)
                    
                    # Check if the link text matches any of the desired site links
                    if link_text_pattern.search(link_text):
                        # Check if the link text matches any meeting/event-related words
                        if any(word.lower() in link_text.lower() for word in relevant_meetings_words):
                            site_links.append((link_url, "Meetings & Events"))
                        elif any(word.lower() in link_text.lower() for word in relevant_Entertainment_words):
                            site_links.append((link_url, "Entertainment"))
                        elif any(word.lower() in link_text.lower() for word in relevant_Facilities_Activities_words):
                            site_links.append((link_url, "Facilities & Activities"))
                        elif any(word.lower() in link_text.lower() for word in relevant_Spa_Wellness_words):
                            site_links.append((link_url, "Spa & Wellness"))
                        elif any(word.lower() in link_text.lower() for word in relevant_Photo_Gallery_words):
                            site_links.append((link_url, "Photo Gallery"))
                        elif any(word.lower() in link_text.lower() for word in relevant_Dining_words):
                            site_links.append((link_url, "Dining"))
                        elif any(word.lower() in link_text.lower() for word in relevant_Location_words):
                            site_links.append((link_url, "Location"))
                        elif any(word.lower() in link_text.lower() for word in relevant_Rooms_words):
                            site_links.append((link_url, "Rooms & Suites"))
                        elif any(word.lower() in link_text.lower() for word in relevant_special_offer_words):
                            site_links.append((link_url, "Special Offers"))
                        elif any(word.lower() in link_text.lower() for word in relevant_Accommodation_words):
                            site_links.append((link_url, "Accommodation"))
                        elif any(word.lower() in link_text.lower() for word in relevant_specials_packages_words):
                            site_links.append((link_url, "Specials & Packages"))
                        else:
                            # Append both link URL and link text
                            site_links.append((link_url, link_text))

                        # Break the loop if the maximum number of links is reached
                        if len(site_links) >= max_links:
                            break
                    else:
                        print(f"Skipping irrelevant link: {link_text} ({link_url})")
                else:
                    print(f"Duplicate link detected: {link_url}")
            else:
                print(f"Skipping invalid link: {link_url}")

        return site_links

    except Exception as e:
        print("An error occurred while scraping the site links:", e)
        return None

def scrape_similar_hotels(google_url, header_text):
    
    try:
        
        options = Options()
        options.add_argument('--no-sandbox')
        options.add_argument('--window-size=1420,1080')
        options.add_argument('--headless')
        options.add_argument('--disable-gpu')
        chrome_driver_path="C:\\Users\\Jyothi.S\\AppData\\Local\\chromedriver-win64\\chromedriver.exe"

        service = Service(chrome_driver_path)
        driver = webdriver.Chrome(service=service, options=options)
        driver.get(google_url)
        time.sleep(5)
 
        search_box = driver.find_element(By.XPATH, "//textarea[@id='APjFqb' and @name='q']")
        search_box.send_keys(header_text)
        search_box.send_keys(Keys.RETURN)
        time.sleep(10)
 
        search_results = driver.find_elements(By.XPATH, "//div[@class='hrZZ8d']")
 
        negative_keywords = []
        for result in search_results:
            negative_keywords.append(result.text)
           
         # Debug print to check the initial negative keywords
        print("Negative Keywords before filtering:", negative_keywords)    
           
        # # Remove 'hotel' and 'resort' from the keywords
        # filtered_negative_keywords  = [keyword.replace(' Hotel', '').replace('Resort', '').strip() for keyword in negative_keywords]
       
        # Debug print to check the filtered negative keywords
        # print("Negative Keywords after filtering:", filtered_negative_keywords)

        # Split the header text into individual words
        header_words = set(header_text.lower().split())

        # Filter out "Hotel" and "Resort" from the header words
        filtered_header_words = {word for word in header_words if word not in ['hotel', 'resort']}

        # Remove keywords containing words from filtered header words
        filtered_negative_keywords= [keyword for keyword in negative_keywords if not any(word in keyword.lower().split() for word in filtered_header_words)]
        print("filtered_negative_keywords",filtered_negative_keywords)
        final_negative_keywords = []
        for keyword in filtered_negative_keywords:
            # Use regex to remove 'hotel' and 'resort' in a case-insensitive manner
            clean_keyword = re.sub(r'\b(hotel|resort)\b', '', keyword, flags=re.IGNORECASE).strip()
            final_negative_keywords.append(clean_keyword)

        print("final_negative_keywords:", final_negative_keywords)
        # Remove 'hotel' and 'resort' from the keywords
        # final_negative_keywords  = [keyword.replace(' Hotel', '').replace('Resort', '').strip() for keyword in filtered_negative_keywords]
       

        # print("final_negative_keywords",final_negative_keywords)

   
 
        # Close the browser
        driver.quit()
 
        # print("Negative Keywords:", negative_keywords)
        return final_negative_keywords
 
    except Exception as e:
        print("An error occurred while scraping similar hotels:", e)
        return None
   
# Define the categorized amenities
amenities_to_check = {
    "Swimming Pool": ["Swimming Pool", "Poolside", "Pool area", "Pool deck", "Pool bar"],
    "Beach Access": ["Beach Access"],
    "Spa Services": ["Spa Services"],
    "Gourmet Dining": ["Gourmet Dining"],
    "Free Breakfast": ["Free Breakfast","Breakfast"],
    "Free Parking": ["Free Parking","Separate parking"],
    "Fitness Center": ["Fitness Center","Fitness Space"],
    "Room Service": ["Room Service"],
    "Daily Housekeeping": ["Daily Housekeeping"],
    "Free WiFi": ["Free WiFi", "Public Wi-Fi", "Wi-Fi Internet Access", "Wi-Fi"],
    "Business Center": ["Business Center"],
    "Air Conditioning": ["A/C", "Air-conditioning", "Air Conditioning & Heating", "Air Conditioning"],
    "Laundry Services": ["Laundry Services","Outsourced Laundry"],
    "Easy Check In/Out": ["Easy Check In", "Express Check Out"],
    "Phone": ["Phone"],
    "Hair Dryer": ["Hair Dryer"],
    "Bicycle Rental": ["Bicycle Rental"],
    "Balcony": ["Balcony", "Balcony/terrace"],
    "Lift": ["Lift"],
    "Iron & Ironing Board": ["Iron & Ironing Board"]
}

# Define a custom exception for timeout
class TimeoutException(Exception):
    pass

def scrape_amenities(url):
    try:
        # Check if the URL is a tel: or mailto: link
        if url.startswith(('tel:', 'mailto:')):
            print("Skipping URL:", url)
            return []
 
        # Fetch the HTML content of the webpage with a timeout
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()  # Raise an exception for non-HTTP or non-HTTPS URLs
 
        # Parse the HTML content
        soup = BeautifulSoup(response.text, 'html.parser')
        all_text = soup.get_text()
 
        # Find amenities
        found_amenities = []
        for amenity, keywords in amenities_to_check.items():
            for keyword in keywords:
                if re.search(re.escape(keyword), all_text, re.IGNORECASE):
                    if amenity not in found_amenities:
                        found_amenities.append(amenity)
       
        print("found_amenities:", found_amenities)
        return found_amenities[:8]
    except Exception as e:
        print(f"An error occurred while scraping amenities from url {url}: {e}")
        return []

def fetch_amenities_from_links(site_links):
    amenities_found = []
    for link_url, _ in site_links:
        try:
            amenities = scrape_amenities(link_url)
            if amenities:
                amenities_found.extend(amenities)
        except Exception as e:
            print(f"An error occurred while fetching amenities from link_url {link_url}: {e}")
    return amenities_found[:8]

def fetch_amenities_from_sub_links(site_links, max_sub_links=20, timeout=6, depth=1):
    amenities_found = set()
    
    def scrape_links(link_url, current_depth):
        nonlocal amenities_found
        try:
            response = requests.get(link_url, headers=headers, timeout=timeout)
            response.raise_for_status()
            amenities = scrape_amenities(link_url)
            if amenities:
                amenities_found.update(amenities)
 
            if current_depth < depth:
                soup = BeautifulSoup(response.text, 'html.parser')
                anchor_tags = soup.find_all('a', href=True)
                unique_urls = set()
                sub_links = []
 
                for a in anchor_tags:
                    sub_link_url = a['href']
                    sub_link_url = urljoin(link_url, sub_link_url)
                    if sub_link_url not in unique_urls:
                        unique_urls.add(sub_link_url)
                        sub_links.append(sub_link_url)
                        if len(sub_links) >= max_sub_links:
                            break
 
                for sub_link_url in sub_links:
                    scrape_links(sub_link_url, current_depth + 1)
 
        except requests.Timeout:
            print(f"Timeout occurred while fetching amenities from sub-link: {link_url}")
        except Exception as e:
            print(f"An error occurred while fetching amenities from sub-link {link_url}: {e}")
 
    for link_url, _ in site_links:
        scrape_links(link_url, 1)
 
    return list(amenities_found)[:8]



if st.session_state.logged_in:
    # Render SEM Creation Template or other content
    st.title("SEM Creation Template")
    url = st.text_input("Enter URL")
    output_file = st.text_input("Enter Header")
    depth = st.number_input("Enter depth", min_value=1, step=1)

    if st.button("Scrape Data"):
        if url:
            try:
                # Initialize variables
                ad_copy1, ad_copy2 = None, None
                header_text = None
                amenities_found = []
                site_links = []
                amenities_from_links = []
                amenities_from_sub_links = []
            # Assuming these functions are defined elsewhere
                try:
                    ad_copy1, ad_copy2 = scrape_first_proper_paragraph(url)
                except:''
        
                header_text = extract_header_from_path(output_file) if output_file else None
        
                amenities_found = scrape_amenities(url)
                print("amenities_found", amenities_found)
        
                # Fetch amenities from link URLs
                site_links = scrape_site_links(url)
                if site_links:
                    amenities_from_links = fetch_amenities_from_links(site_links)
                else:
                    print("No site links found.")
                    amenities_from_links = []
                print("amenities_from_links", amenities_from_links)
        
                # Fetch amenities from subsequent links with specified depth
                amenities_from_sub_links = fetch_amenities_from_sub_links(site_links, max_sub_links=20, depth=depth)
                print("amenities_from_sub_links", amenities_from_sub_links)
        
                # Combine all fetched amenities
                all_amenities = amenities_found + amenities_from_links + amenities_from_sub_links
                unique_amenities = list(set(all_amenities))[:8]
            
                sub_links_processed = 0
                while 4 < len(unique_amenities) < 8 and len(site_links) > 0 and sub_links_processed < 20:
                    max_sub_links = 20 - sub_links_processed  # Fetch amenities from remaining sub-links
                    additional_amenities_from_sub_links = fetch_amenities_from_sub_links(site_links, max_sub_links)
                    unique_amenities.extend(additional_amenities_from_sub_links)
                    unique_amenities = list(set(unique_amenities))[:8]  # Limit to a maximum of 8 unique amenities
                    sub_links_processed += max_sub_links  # Update the number of sub-links processed
                    if sub_links_processed >= 20:
                        break  # Break out of the loop after checking 20 sub-links
        
                amenity_order = list(amenities_to_check.keys())
                sorted_amenities = sorted(unique_amenities, key=lambda x: amenity_order.index(x) if x in amenity_order else len(amenity_order))


                property_name_variants = generate_variants(header_text) if header_text else []

                final_negative_keyword  = scrape_similar_hotels("https://www.google.com", header_text) if header_text else []
                
                # Debug print to check the final filtered negative keywords
                print("Final filtered Negative Keywords:", final_negative_keyword)

                header_df = pd.DataFrame({'Header Text': [header_text] if header_text else []})
                paragraph_df = pd.DataFrame({'Ad copy1': [ad_copy1], 'Ad copy2': [ad_copy2]})
                site_links_df = pd.DataFrame(site_links, columns=['Link URL', 'Link Text'])
                property_url_df = pd.DataFrame({'property_url': [url]})
                property_name_variants_df = pd.DataFrame({'Variants of Property Name': property_name_variants})
                negative_keywords_df = pd.DataFrame(final_negative_keyword, columns=['Negative Keywords'])
                amenities_df = pd.DataFrame({'Amenities': sorted_amenities})
                Callouts = ["Book Direct", "Great Location", "Spacious Suites"]

                df = pd.concat([header_df, paragraph_df, site_links_df, property_url_df, property_name_variants_df, negative_keywords_df, amenities_df], axis=1)
                print("first dataframe",df)
                try:
                    response = requests.get(url, timeout=10)
                    response.raise_for_status()
                    page_content = response.text
                    water_keywords = ["swimming pool", "Water Park", "pool", "sea", "Salt Water Swimming Pool", "Pool & sea", "Poolside", "Pool area", "Pool deck", "Pool bar"]
                    balcony_keywords = ["balcony", "terrace", "veranda", "patio", "deck", "outdoor seating", "private balcony", "balcony view", "balcony access", "sun deck", "rooftop terrace", "lanai", "courtyard", "loggia", "open-air balcony", "French balcony", "wrap-around balcony", "overlooking balcony", "scenic balcony", "balcony suite"]
                    pet_keywords = ["pet-friendly","Pet Friendly", "dog friendly","dog","pet-friendly policy", "dog-friendly", "cat-friendly", "pet-friendly hotel", "pet-friendly apartment", "pet-friendly rental", "pet-friendly room", "pet-friendly amenities", "pet-friendly patio", "pet-friendly park", "pet-friendly restaurant", "pet-friendly neighborhood", "pet-friendly community", "pet-friendly activities", "pet-friendly events", "pet-friendly travel", "pet-friendly vacations", "pet-friendly establishments"]

                    water_found = [keyword for keyword in water_keywords if re.search(keyword, page_content, re.IGNORECASE)]
                    balcony_found = [keyword for keyword in balcony_keywords if re.search(keyword, page_content, re.IGNORECASE)]
                    pet_found = [keyword for keyword in pet_keywords if re.search(keyword, page_content, re.IGNORECASE)]

                    print("Water-related keywords found:", water_found)
                    print("Balcony-related keywords found:", balcony_found)
                    print("Pet-friendly keywords found:", pet_found)

                    if any(re.search(keyword, page_content, re.IGNORECASE) for keyword in water_keywords):
                        Callouts.append("Water Park")
                    if any(re.search(keyword, page_content, re.IGNORECASE) for keyword in balcony_keywords):
                        Callouts.append("Balcony")
                    if any(re.search(keyword, page_content, re.IGNORECASE) for keyword in pet_keywords):
                        Callouts.append("Pet-friendly")
                except Exception as e:
                    print(f"An error occurred while checking for water-related keywords: {e}")

                callouts_df = pd.DataFrame({'Callouts': Callouts})
                df = pd.concat([df, callouts_df], axis=1)
                
                st.write("Final DataFrame:", df)

                def save_df_to_excel(df):
                    output = BytesIO()
                    writer = pd.ExcelWriter(output, engine='openpyxl')
                    
                    df.to_excel(writer, index=False, sheet_name='Sheet1')
                    workbook = writer.book
                    worksheet = writer.sheets['Sheet1']

                    # Define header styles
                    header_fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
                    header_font = Font(color='FFFFFF', bold=True)
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font

                    # Adjust column widths
                    padding = 5
                    specific_columns = {
                        0: 15, 1: 15, 2: 15, 3: 40, 4: 18, 5: 30, 6: 22, 7: 30, 8: 18, 9: 15
                    }
                    for i, column in enumerate(worksheet.columns):
                        if i in specific_columns:
                            adjusted_width = specific_columns[i]
                        else:
                            max_length = 0
                            column = [cell for cell in column]
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(cell.value)
                                except:
                                    pass
                            adjusted_width = (max_length + padding)
                        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

                    writer._save()
                    processed_data = output.getvalue()
                    return processed_data
                

                # Save DataFrame to Excel and get the byte data
                excel_data = save_df_to_excel(df)

                # Create a download button
                st.download_button(
                    label="Download Excel file",
                    data=excel_data,
                    file_name="formatted_data.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"Error occurred: {e}")

        else:
            st.warning("Please enter a URL.")

    # Logout button
    if st.button("Logout"):
        logout()      

else:
    # Render login page or redirect to login if not logged in
    login_success = login()
    if login_success:
        st.rerun()           


        
